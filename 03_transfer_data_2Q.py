import os
import time
import platform
from typing import List, Optional, Tuple
import openpyxl as px
import polars as pl
import pandas as pd
import warnings
from rich import print
from concurrent.futures import ThreadPoolExecutor, as_completed

# ===== 設定 =====
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.worksheet.header_footer')

# ===== 定数定義 =====
TERM_JP = "第２四半期"
SHEET_NAMES = ["法人", "得意先", "得意先（商品）"]
SOURCE_DIR = "転記用データ"
TEMPLATE_FILE_PATH = "ランキング（得意先）_フォーマット.xlsx"
OUTPUT_DIR = "D:/Skit_Actual/EXCEL/累計/"

def get_proper_path(path: str) -> str:
    """実行環境に応じて適切なパスを返す
    
    Args:
        path: 変換対象のパス（Windows形式 D:/...）
        
    Returns:
        str: 実行環境に適したパス
    """
    if platform.system() == "Linux":  # WSL環境
        return path.replace("D:/", "/mnt/d/")
    else:  # Windows環境
        return path

def validate_files_exist(source_dir: str, template_path: str, sheet_names: List[str]) -> Tuple[bool, List[str]]:
    """ファイル存在確認（戻り値を改善）
    
    Returns:
        Tuple[bool, List[str]]: (全ファイル存在フラグ, 存在するparquetファイルのリスト)
    """
    if not os.path.exists(template_path):
        print(f"[red]エラー: テンプレートファイル '{template_path}' が存在しません[/red]")
        return False, []
    
    if not os.path.exists(source_dir):
        print(f"[red]エラー: ソースディレクトリ '{source_dir}' が存在しません[/red]")
        return False, []
    
    existing_files = []
    missing_files = []
    
    for sheet_name in sheet_names:
        parquet_file = os.path.join(source_dir, f"{sheet_name}_転記用データ.parquet")
        if os.path.exists(parquet_file):
            existing_files.append(parquet_file)
        else:
            missing_files.append(parquet_file)
    
    if missing_files:
        print(f"[yellow]警告: 以下のparquetファイルが存在しません:[/yellow]")
        for file in missing_files:
            print(f"  - {file}")
        print(f"[yellow]利用可能なファイル: {len(existing_files)}個[/yellow]")
    
    return len(missing_files) == 0, existing_files

def clear_worksheet_data(worksheet) -> None:
    if worksheet.max_row > 0:
        worksheet.delete_rows(1, worksheet.max_row)

def write_dataframe_to_worksheet(df: pl.DataFrame, worksheet) -> None:
    """polarsデータフレームをワークシートに書き込み（最適化版）"""
    # ヘッダー行を追加
    worksheet.append(df.columns)
    
    # データ行を一括で追加（polarsの高速イテレーション）
    for row in df.iter_rows():
        worksheet.append(list(row))

def process_single_sheet(workbook, source_dir: str, sheet_name: str) -> Tuple[bool, str]:
    """単一シートの処理（エラーハンドリング強化版）
    
    Returns:
        Tuple[bool, str]: (成功フラグ, 結果メッセージ)
    """
    parquet_file = os.path.join(source_dir, f"{sheet_name}_転記用データ.parquet")
    
    if not os.path.exists(parquet_file):
        msg = f"ソースファイル '{parquet_file}' が存在しません"
        print(f"[yellow]警告: {msg}。スキップします。[/yellow]")
        return False, msg
    
    if sheet_name not in workbook.sheetnames:
        msg = f"テンプレートファイルにシート '{sheet_name}' が存在しません"
        print(f"[yellow]警告: {msg}。スキップします。[/yellow]")
        return False, msg
    
    try:
        print(f"[cyan]シート '{sheet_name}' のデータを処理中...[/cyan]")
        
        # polarsで高速読み込み
        df = pl.read_parquet(parquet_file)
        
        # ワークシートのクリアと書き込み
        worksheet = workbook[sheet_name]
        clear_worksheet_data(worksheet)
        write_dataframe_to_worksheet(df, worksheet)
        
        msg = f"シート '{sheet_name}' の処理が完了 (行数: {len(df)})"
        print(f"[green]{msg}[/green]")
        return True, msg
        
    except Exception as e:
        msg = f"シート '{sheet_name}' の処理中にエラー: {e}"
        print(f"[red]{msg}[/red]")
        return False, msg

def transfer_data(source_dir: str, template_path: str, output_path: str, sheet_names: List[str]) -> Tuple[bool, List[str]]:
    """データ転記処理（エラーハンドリング強化版）
    
    Returns:
        Tuple[bool, List[str]]: (成功フラグ, 処理結果メッセージのリスト)
    """
    workbook: Optional[px.Workbook] = None
    results = []
    
    try:
        print(f"[cyan]テンプレートファイルを読み込み中: {template_path}[/cyan]")
        workbook = px.load_workbook(template_path)
        
        success_count = 0
        
        # 注意: openpyxlはスレッドセーフではないため、逐次処理に変更
        print(f"[yellow]シート処理を開始（openpyxlの制約により逐次処理）[/yellow]")
        
        for sheet_name in sheet_names:
            success, msg = process_single_sheet(workbook, source_dir, sheet_name)
            results.append(msg)
            if success:
                success_count += 1
        
        if success_count > 0:
            print(f"[cyan]Excelファイル保存中: {output_path}[/cyan]")
            
            # 出力ディレクトリが存在しない場合は作成
            output_dir = os.path.dirname(output_path)
            # 環境に応じたパス変換
            actual_output_dir = get_proper_path(output_dir)
            os.makedirs(actual_output_dir, exist_ok=True)
            
            # 環境に応じたパス変換
            actual_output_path = get_proper_path(output_path)
            workbook.save(actual_output_path)
            
            final_msg = f"データ転記が完了しました ({success_count}/{len(sheet_names)} シート処理完了)"
            print(f"[green]{final_msg}[/green]")
            print(f"[green]保存先: {output_path}[/green]")
            print(f"[green]実際のパス: {actual_output_path}[/green]")
            results.append(final_msg)
            return True, results
        else:
            error_msg = "処理できるシートがありませんでした"
            print(f"[red]{error_msg}[/red]")
            results.append(error_msg)
            return False, results
            
    except FileNotFoundError as e:
        error_msg = f"ファイルが見つかりません - {e}"
        print(f"[red]エラー: {error_msg}[/red]")
        results.append(error_msg)
        return False, results
    except PermissionError as e:
        error_msg = f"ファイルへのアクセス権限がありません - {e}"
        print(f"[red]エラー: {error_msg}[/red]")
        results.append(error_msg)
        return False, results
    except Exception as e:
        error_msg = f"予期せぬエラーが発生しました: {e} (エラータイプ: {type(e).__name__})"
        print(f"[red]{error_msg}[/red]")
        results.append(error_msg)
        return False, results
    finally:
        if workbook:
            workbook.close()

def main() -> None:
    print(f"[bold cyan]データ転記処理を開始します（polars最適化版）[/bold cyan]")
    print(f"対象期間: {TERM_JP}")
    print()
    
    start_time = time.time()
    
    save_file_name = f"{TERM_JP}_ランキング（得意先）.xlsx"
    save_path = os.path.join(OUTPUT_DIR, save_file_name)
    
    print(f"[bold]処理設定:[/bold]")
    print(f"  ソースディレクトリ: {SOURCE_DIR}")
    print(f"  テンプレートファイル: {TEMPLATE_FILE_PATH}")
    print(f"  出力パス: {save_path}")
    print(f"  処理対象シート: {', '.join(SHEET_NAMES)}")
    print()
    
    # ファイル存在確認（改良版）
    all_exists, existing_files = validate_files_exist(SOURCE_DIR, TEMPLATE_FILE_PATH, SHEET_NAMES)
    
    if not existing_files:
        print(f"[red]利用可能なファイルがありません。処理を中止します。[/red]")
        return
    elif not all_exists:
        print(f"[yellow]一部のファイルが不足していますが、利用可能なファイルで処理を続行します。[/yellow]")
    
    # 利用可能なシート名を抽出
    available_sheets = []
    for file_path in existing_files:
        sheet_name = os.path.basename(file_path).replace("_転記用データ.parquet", "")
        available_sheets.append(sheet_name)
    
    print(f"[cyan]処理対象シート: {', '.join(available_sheets)}[/cyan]")
    
    # データ転記実行
    success, messages = transfer_data(SOURCE_DIR, TEMPLATE_FILE_PATH, save_path, available_sheets)
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    
    print()
    print(f"[bold]処理結果:[/bold]")
    print(f"  実行時間: {elapsed_time:.2f}秒")
    print(f"  処理状況: {'[green]成功[/green]' if success else '[red]失敗[/red]'}")
    print(f"  使用ライブラリ: polars (高速parquet読み込み)")
    print(f"  エラーハンドリング: 強化済み")
    
    # 詳細メッセージ表示
    if messages:
        print(f"\n[bold]詳細ログ:[/bold]")
        for i, msg in enumerate(messages, 1):
            print(f"  {i}. {msg}")

if __name__ == "__main__":
    main()